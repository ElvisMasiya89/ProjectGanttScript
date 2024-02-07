import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Define the tasks, their start and end dates
tasks_data = [
    {"Task": "Literature Review", "Start": "2024-02-02", "End": "2024-03-03"},
    {"Task": "Experiment Design", "Start": "2024-02-02", "End": "2024-02-16"},
    {"Task": "Data Preparation", "Start": "2024-02-17", "End": "2024-03-08"},
    {"Task": "Model Selection and Setup", "Start": "2024-02-27", "End": "2024-03-08"},
    {"Task": "Run Experiments", "Start": "2024-03-09", "End": "2024-04-07"},
    {"Task": "Prepare and Conduct Survey", "Start": "2024-03-09", "End": "2024-04-07"},
    {"Task": "Data Analysis", "Start": "2024-04-08", "End": "2024-04-22"},
    {"Task": "Analyze Survey Results", "Start": "2024-04-08", "End": "2024-04-22"},
    {"Task": "Writing - Draft", "Start": "2024-04-08", "End": "2024-04-27"},
    {"Task": "Revision and Finalization", "Start": "2024-04-28", "End": "2024-04-30"},
    {"Task": "Incorporate Survey Findings", "Start": "2024-04-23", "End": "2024-04-30"},
]

# Convert to DataFrame
tasks = pd.DataFrame(tasks_data)

# Convert start and end dates to datetime
tasks['Start'] = pd.to_datetime(tasks['Start'])
tasks['End'] = pd.to_datetime(tasks['End'])
tasks['Duration'] = tasks['End'] - tasks['Start']

# Sort tasks by start date
tasks = tasks.sort_values(by='Start')

# Plotting
fig, ax = plt.subplots(figsize=(8, 6))  # Adjusted for smaller size

# Create bars for tasks
for i, task in tasks.iterrows():
    ax.barh(task['Task'], task['Duration'], left=task['Start'], color='skyblue', edgecolor='black')

# Format the date axis
ax.xaxis_date()
ax.xaxis.set_major_locator(mdates.WeekdayLocator())
ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))

plt.xticks(rotation=45)
plt.xlabel('Date')
plt.ylabel('Tasks')
plt.title('Project Timeline Gantt Chart')

# Save the chart as an image
chart_filename = 'gantt_chart.png'
plt.tight_layout()
plt.savefig(chart_filename)
plt.close()

# Write tasks to an Excel file
excel_filename = 'project_timeline.xlsx'
tasks.to_excel(excel_filename, index=False)

# Load the Excel file using openpyxl
book = load_workbook(excel_filename)

# Get the first sheet of the workbook
sheet = book.active

# Calculate insertion point to the right of the table
num_tasks = len(tasks)
start_row_image = 2
start_col_image = len(tasks.columns) + 2  # Adding some padding to the right

# Insert the Gantt chart image
img = Image(chart_filename)
# Place the image to the right of the table
cell_ref = f'{get_column_letter(start_col_image)}{start_row_image}'
sheet.add_image(img, cell_ref)

# Save the modified Excel file
book.save(excel_filename)
